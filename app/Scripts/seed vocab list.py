from openpyxl import load_workbook
# import database model
from app import db

def seeder():
# Take all the rows in all sheet in the Vocab List and make tile model for each row


  #Load the excel file
  workbook=load_workbook(filename='')
  
  #loop the sheets in the workbook
  for sheet_name in workbook.sheetnames:
    sheet=workbook[sheet_name]
    
    #loop rows in the sheet
    for row in sheet.iter_rows(min_row=2):
      #create a title model for each row
      tile=Tile(word=row[0].value,definition=row[1].value)
      
      #add the title model to the database
      db.session.add(title)
      
  #commit the chnages to the database
  db.session.commit()
      
